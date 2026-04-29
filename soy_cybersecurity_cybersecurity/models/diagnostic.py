# -*- coding: utf-8 -*-

import base64
import os
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile

from odoo import api, fields, models
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from odoo.exceptions import UserError
import logging 

_logger = logging.getLogger(__name__)

# Porcentajes para las barras de cargas
AVAILABLE_PRIORITIES = [
    ('na', 'N/A - No aplica'),
    ('0_porcent', '0% - No documentado / No existente'),
    ('25_porcent', '25% - Aplicado / No documentado'),
    ('50_porcent', '50% - Documentado / No aplicado'),
    ('75_porcent', '75% - Aplicado y documentado'),
    ('100_porcent', '100% - Aplicado, documentado y controlado')]

FIELDS = ( 
    'diagnostic4_1_ids', 'diagnostic4_2_ids', 'diagnostic4_3_ids', 'diagnostic4_4_ids',
    'diagnostic5_1_ids', 'diagnostic5_2_ids', 'diagnostic5_3_ids', 
    'diagnostic6_1_ids', 'diagnostic6_2_ids',
    'diagnostic7_1_ids', 'diagnostic7_2_ids', 'diagnostic7_3_ids', 'diagnostic7_4_ids', 'diagnostic7_5_ids', 
    'diagnostic8_1_ids', 'diagnostic8_2_ids',  'diagnostic8_3_ids', 
    'diagnostic9_1_ids', 'diagnostic9_2_ids', 'diagnostic9_3_ids',
          'diagnostic10_1_ids', 'diagnostic10_2_ids', )

class Requirement(models.Model):
    _name = 'cybersecurity.diagnostic.requirement'
    _description = "Requirimientos de diagnostico de SGSI"

    name = fields.Char(string=u'Nombre', required=True)
    complete_name = fields.Text(string=u'Descripción', required=True)
    info = fields.Text(string=u'Interpretación', store=True)
    clause_id = fields.Many2one(
        string=u'Clausula', comodel_name='cybersecurity.clause', required=True)
    chapter = fields.Selection(
        string=u'Capítulo', related='clause_id.chapter', store=True)
    position_excel = fields.Char(string=u'Posición en excel')


class DiagnosticLine(models.Model):
    _name = 'cybersecurity.diagnostic.line'
    _description = 'Linea de análisis de calidad'

    requirement_name = fields.Char(related='requirement_id.name')
    # FIX
    diagnostic4_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic4_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic4_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic4_4_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='cybersecurity.diagnostic', ondelete='cascade')

    diagnostic5_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic5_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic5_3_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    
    diagnostic6_1_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic6_2_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    
    diagnostic7_1_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic7_2_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic7_3_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic7_4_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic7_5_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    
    diagnostic8_1_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic8_2_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic8_3_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
        
    diagnostic9_1_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic9_2_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic9_3_id = fields.Many2one(
        string=u'Diagnostico', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    
    diagnostic10_1_id = fields.Many2one(
        string=u'Diagnostico 10.1', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    diagnostic10_2_id = fields.Many2one(
        string=u'Diagnostico 10.2', 
        comodel_name='cybersecurity.diagnostic', ondelete='cascade')
    
    requirement_id = fields.Many2one(
        string=u'Requisito',
        comodel_name='cybersecurity.diagnostic.requirement',)
    clause_id = fields.Text(string=u'Clausula ID', store=True)
    clause = fields.Many2one(
        string=u'Claúsulas', 
        comodel_name='cybersecurity.clause',
        ondelete='cascade')

    info = fields.Text(string=u'Interpretación',
                       help="here is my message", store=True)

    name = fields.Text(string=u'Nombre requirimiento', store=True)

    qualification = fields.Selection(AVAILABLE_PRIORITIES,
                                     index=True,
                                     string=u'Calificación',
                                     required=True,
                                     default='na',
                                     store=True)

    observation = fields.Text(string=u'Observaciones')
    is_page = fields.Boolean('Is a page?')
    display_type = fields.Selection([
        ('line_section', 'Section'),
        ('line_note', 'Note'),
    ], default=False, help="Technical field for UX purpose.")

    
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('name') and vals.get('requirement_id'):
                requirement = self.env['cybersecurity.diagnostic.requirement'].browse(vals['requirement_id'])
                vals['name'] = requirement.complete_name

        records = super().create(vals_list)
        return records

    @api.onchange('requirement_id')
    def _onchange_requirement_id(self):
        self.name = self.requirement_id.complete_name


class Clause(models.Model):
    _name = 'cybersecurity.clause'
    _description = "Claúsulas"

    question = fields.Text(string=u'Pregunta ref.', required=True)
    name = fields.Char(string=u'Nombre', required=True)
    complete_name = fields.Text(string=u'Descripción', required=True)
    chapter = fields.Selection(
        string=u'Capítulo',
        selection=[
            ('4_context', 'Contexto de la organización'),
            ('5_leadership', 'Liderazgo'),
            ('6_planning', 'Planificación'),
            ('7_support', 'Apoyo'),
            ('8_operation', 'Operación'),
            ('9_evaluation', 'Evaluación del desempeño'),
            ('10_improvement', 'Mejora')],
        required=True,
    )


class Diagnostic(models.Model):
    _name = 'cybersecurity.diagnostic'
    _description = "Diagnostico SI"

    name = fields.Char(string=u'Nombre', required=True,
                       default=lambda self: "Análisis de ...")
    user_id = fields.Many2one(
        string='Responsable',
        comodel_name='res.users',
        ondelete='cascade',
        default=lambda self: self.env.user and self.env.user.id or False,
    )

    company_id = fields.Many2one(
        string=u'Compañia', 
        comodel_name='res.company', 
        required=True,
        domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)], 
        default=lambda self: self.env.company,
    )
    date_diagnostic = fields.Datetime(
        string=u'Fecha creación', default=fields.Datetime.now, required=True)
    date_validate = fields.Datetime(
        string=u'Fecha evaluación', related='xls_helper.date_validate')

    all_clause = fields.Many2many(
        comodel_name='cybersecurity.clause', string=u'Clausulas')


    diagnostic4_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic4_1_id',)
    diagnostic4_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic4_2_id',)
    diagnostic4_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic4_3_id',)
    diagnostic4_4_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic4_4_id',)

    diagnostic5_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic5_1_id',)
    diagnostic5_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic5_2_id',)
    diagnostic5_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic5_3_id',)

    diagnostic6_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic6_1_id',)
    diagnostic6_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic6_2_id',)

    diagnostic7_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic7_1_id',)
    diagnostic7_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic7_2_id',)
    diagnostic7_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic7_3_id',)
    diagnostic7_4_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic7_4_id',)
    diagnostic7_5_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic7_5_id',)

    diagnostic8_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic8_1_id', )
    diagnostic8_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic8_2_id', )
    diagnostic8_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic8_3_id', )

    diagnostic9_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic9_1_id', )
    diagnostic9_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic9_2_id', )
    diagnostic9_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic9_3_id', )
    

    diagnostic10_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic10_1_id', )
    diagnostic10_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='cybersecurity.diagnostic.line', inverse_name='diagnostic10_2_id', )

    state = fields.Selection(
        string=u'Estado',
        selection=[('draft', 'Previo'),
                   ('evaluate', 'Detallado'),
                   ('validate', 'Culminado')],
        default='draft',
    )

    diagnostic4_ids_100 = fields.Integer(
        string=u'Total Contexto 100%', compute='_get_diagnostic', store=True)
    diagnostic4_ids_75 = fields.Integer(
        string=u'Total Contexto 75%', compute='_get_diagnostic', store=True)
    diagnostic4_ids_50 = fields.Integer(
        string=u'Total Contexto 50%', compute='_get_diagnostic', store=True)
    diagnostic4_ids_25 = fields.Integer(
        string=u'Total Contexto 25%', compute='_get_diagnostic', store=True)
    diagnostic4_ids_0 = fields.Integer(
        string=u'Total Contexto 0%', compute='_get_diagnostic', store=True)

    diagnostic5_ids_100 = fields.Integer(
        string=u'Total Liderazgo 100%', compute='_get_diagnostic', store=True)
    diagnostic5_ids_75 = fields.Integer(
        string=u'Total Liderazgo 75%', compute='_get_diagnostic', store=True)
    diagnostic5_ids_50 = fields.Integer(
        string=u'Total Liderazgo 50%', compute='_get_diagnostic', store=True)
    diagnostic5_ids_25 = fields.Integer(
        string=u'Total Liderazgo 25%', compute='_get_diagnostic', store=True)
    diagnostic5_ids_0 = fields.Integer(
        string=u'Total Liderazgo 0%', compute='_get_diagnostic', store=True)

    diagnostic6_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic', store=True)
    diagnostic6_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic', store=True)
    diagnostic6_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic', store=True)
    diagnostic6_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic', store=True)
    diagnostic6_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic', store=True)

    diagnostic7_ids_100 = fields.Integer(
        string=u'Total Apoyo 100%', compute='_get_diagnostic', store=True)
    diagnostic7_ids_75 = fields.Integer(
        string=u'Total Apoyo 75%', compute='_get_diagnostic', store=True)
    diagnostic7_ids_50 = fields.Integer(
        string=u'Total Apoyo 50%', compute='_get_diagnostic', store=True)
    diagnostic7_ids_25 = fields.Integer(
        string=u'Total Apoyo 25%', compute='_get_diagnostic', store=True)
    diagnostic7_ids_0 = fields.Integer(
        string=u'Total Apoyo 0%', compute='_get_diagnostic', store=True)

    diagnostic8_ids_100 = fields.Integer(
        string=u'Total Operación 100%', compute='_get_diagnostic', store=True)
    diagnostic8_ids_75 = fields.Integer(
        string=u'Total Operación 75%', compute='_get_diagnostic', store=True)
    diagnostic8_ids_50 = fields.Integer(
        string=u'Total Operación 50%', compute='_get_diagnostic', store=True)
    diagnostic8_ids_25 = fields.Integer(
        string=u'Total Operación 25%', compute='_get_diagnostic', store=True)
    diagnostic8_ids_0 = fields.Integer(
        string=u'Total Operación 0%', compute='_get_diagnostic', store=True)

    diagnostic9_ids_100 = fields.Integer(
        string=u'Total Desempeño 100%', compute='_get_diagnostic', store=True)
    diagnostic9_ids_75 = fields.Integer(
        string=u'Total Desempeño 75%', compute='_get_diagnostic', store=True)
    diagnostic9_ids_50 = fields.Integer(
        string=u'Total Desempeño 50%', compute='_get_diagnostic', store=True)
    diagnostic9_ids_25 = fields.Integer(
        string=u'Total Desempeño 25%', compute='_get_diagnostic', store=True)
    diagnostic9_ids_0 = fields.Integer(
        string=u'Total Desempeño 0%', compute='_get_diagnostic', store=True)

    diagnostic10_ids_100 = fields.Integer(
        string=u'Total Mejora 100%', compute='_get_diagnostic', store=True)
    diagnostic10_ids_75 = fields.Integer(
        string=u'Total Mejora 75%', compute='_get_diagnostic', store=True)
    diagnostic10_ids_50 = fields.Integer(
        string=u'Total Mejora 50%', compute='_get_diagnostic', store=True)
    diagnostic10_ids_25 = fields.Integer(
        string=u'Total Mejora 25%', compute='_get_diagnostic', store=True)
    diagnostic10_ids_0 = fields.Integer(
        string=u'Total Mejora 0%', compute='_get_diagnostic', store=True)

    diagnostic4_ids_total = fields.Integer(
        string=u'Total Punto 4', compute='_get_diagnostic', store=True)
    diagnostic5_ids_total = fields.Integer(
        string=u'Total Punto 5', compute='_get_diagnostic', store=True)
    diagnostic6_ids_total = fields.Integer(
        string=u'Total Punto 6', compute='_get_diagnostic', store=True)
    diagnostic7_ids_total = fields.Integer(
        string=u'Total Punto 7', compute='_get_diagnostic', store=True)
    diagnostic8_ids_total = fields.Integer(
        string=u'Total Punto 8', compute='_get_diagnostic', store=True)
    diagnostic9_ids_total = fields.Integer(
        string=u'Total Punto 9', compute='_get_diagnostic', store=True)
    diagnostic10_ids_total = fields.Integer(
        string=u'Total Punto 10', compute='_get_diagnostic', store=True)

    def get_diagnostic_values(self, diagnostic_list):
        total_100 = total_75 = total_50 = total_25 = total_0 = total_na = 0

        for line in diagnostic_list:
            line_t = getattr(self, line)
            for line_ in line_t:
                if line_.qualification == 'na':
                    total_na += 1
                if line_.qualification == '100_porcent':
                    total_100 += 1
                if line_.qualification == '75_porcent':
                    total_75 += 1
                if line_.qualification == '50_porcent':
                    total_50 += 1
                if line_.qualification == '25_porcent':
                    total_25 += 1
                if line_.qualification == '0_porcent':
                    total_0 += 1
        total_in_partials = [total_na, total_100,
                             total_75, total_50, total_25, total_0]
        sum_total = sum(total_in_partials)
        return [total_na, total_100, total_75, total_50, total_25, total_0, sum_total]

    @api.depends(*FIELDS)
    def _get_diagnostic(self):
        try:
            # Preparar los valores calculados
            field_suffixes = ['4', '5', '6', '7', '8', '9', '10']
            field_string = 'diagnostic{}_ids_{}'
            field_total = 'diagnostic{}_ids_total'

            # Diccionario para almacenar los resultados
            results = {}

            # Hacer uso de sudo() para realizar los cálculos con permisos elevados
            for record in self.sudo():
                for suffix in field_suffixes:
                    diagnostic_list = list(filter(lambda x: x.startswith(f'diagnostic{suffix}'), FIELDS))
                    if diagnostic_list:
                        res = record.get_diagnostic_values(diagnostic_list)
                        results[suffix] = {
                            '100': int(res[1]),
                            '75': int(res[2]),
                            '50': int(res[3]),
                            '25': int(res[4]),
                            '0': int(res[5]),
                            'total': int(res[6]),
                        }

                # Asignar los resultados a los campos fuera del ciclo de cálculo
                updates = {}
                for suffix in field_suffixes:
                    if suffix in results:
                        updates.update({
                            field_string.format(suffix, '100'): results[suffix]['100'],
                            field_string.format(suffix, '75'): results[suffix]['75'],
                            field_string.format(suffix, '50'): results[suffix]['50'],
                            field_string.format(suffix, '25'): results[suffix]['25'],
                            field_string.format(suffix, '0'): results[suffix]['0'],
                            field_total.format(suffix): results[suffix]['total'],
                        })

                # Aplicar los resultados a los campos correspondientes
                record.update(updates)

        except Exception as e:
            raise UserError(f"Ocurrió un error al calcular los diagnósticos: {str(e)}")


    def _default_diagnostic_line_ids(self, vchapter):
        """ 
        """
        requirements = self.env['cybersecurity.diagnostic.requirement'].search(
            [('chapter', '=', vchapter)])

        lines = [(5, 0, 0)]
        for req in requirements:
            if req.clause_id.id == '4.1':
                data = {
                    'name': req.complete_name,
                    'requirement_id': req.id,
                    'clause_id': req.clause_id.id,
                    'qualification': 'na',
                }
                lines.append((0, 0, data))
        return lines

    def _default_diagnostic_line_ids_v2(self, vchapter, ids_clause):
        """ Devuelve los requisitos de las clausulas, de acuerdo
        a su capitulo y id_req
        """
        requirements = self.env['cybersecurity.diagnostic.requirement'].search(
            [('name', '=like', vchapter + '%')])

        lines = [(5, 0, 0)]
        print([x.name for x in requirements])
        for req in requirements:
            if (req.clause_id.id in ids_clause):
                data = {
                    'info': req.info,
                    'name': req.complete_name,
                    'requirement_id': req.id,
                    'clause_id': req.clause_id.id,
                    'qualification': 'na',
                }
                lines.append((0, 0, data))
        return lines

    def _get_xls_helper(self):
        res = self.env['hola_calidad.xls_helper'].search([])
        return res

    datas = fields.Binary('File', readonly=True)
    datas_fname = fields.Char('Filename', readonly=True)
    xls_helper = fields.Many2one(comodel_name='hola_calidad.xls_helper',
                                 string='Soy Calidad', readonly=True, default=_get_xls_helper)

    def evaluate_diagnostic(self):
        paths = os.path.realpath(__file__)
        dirname = os.path.dirname(os.path.dirname(paths))
        newdir = os.path.join(dirname, 'data')
        workbook = load_workbook(newdir+'/data.xlsx')

        # user = self.env['res.users'].browse(self.env.uid)
        user = self.env.user
        time = datetime.now()
        self.xls_helper.write({'date_validate': datetime.now()})
        if time:
            filename = self.name + ' ' + \
                str(time.strftime("%Y-%m-%d %H:%M %p"))
        else:
            filename = self.name

        sheets = workbook.sheetnames
        sheet = workbook[sheets[1]]
        cont = 15
        cont_relle = 0
        MAX_ROW_EXCEL = 126

        for diagnostics in [getattr(self, x) for x in FIELDS]:
            for diagnostic_line in diagnostics:
                p_excel = diagnostic_line.requirement_id.position_excel
                print("p_excel------>", p_excel)
                if p_excel and cont < MAX_ROW_EXCEL:
                    number = p_excel[1:]
                    if number != str(cont):
                        # for i=number; in diagnostics: cell1 = sheet['B'+str(cont)]
                        # if cont+1 in [17, 18,22,23] 31
                        i = cont
                        tmp = 1
                        tmpfinal = True
                        tmp1 = 1
                        while tmpfinal == True and cont < MAX_ROW_EXCEL and i < MAX_ROW_EXCEL:
                            if i not in [14, 19, 18, 24, 23, 27, 28, 32, 31, 33, 37, 38, 43,44, 47, 48,49, 55, 56, 61, 62, 63, 65, 66, 70, 71, 74, 75, 78, 79, 86, 85, 84, 91, 92, 94, 95, 100, 101, 102, 106, 107, 111, 112, 116, 117, 118, 122, 123, ] and i != int(number):
                                if cont < MAX_ROW_EXCEL and i < MAX_ROW_EXCEL:
                                    # comvertir a string
                                    cell1 = sheet['B'+str(i)]
                                    if not isinstance(cell1, MergedCell):
                                        cell1.value = 'X'
                                    tmp = 0
                                    i = i+1
                                    tmp1 = 0
                            else:
                                if tmp == 0:
                                    print("vista tmp y i------>", i)
                                    tmpfinal = False
                                    i = int(number)
                                    break

                            if number == str(i):  # 31
                                tmpfinal = False
                                i = int(number)
                                break
                            else:
                                if tmp1 != 0:
                                    i = i+1
                        cont = i
                    if number == str(cont):
                        number = p_excel[1:]
                        print("number------>", number)
                        if diagnostic_line.qualification == 'na':
                            letter = 'G'
                        elif diagnostic_line.qualification == '0_porcent':
                            letter = 'B'
                        elif diagnostic_line.qualification == '25_porcent':
                            letter = 'C'
                        elif diagnostic_line.qualification == '50_porcent':
                            letter = 'D'
                        elif diagnostic_line.qualification == '75_porcent':
                            letter = 'E'
                        elif diagnostic_line.qualification == '100_porcent':
                            letter = 'F'

                        cell = sheet[letter+number]
                        _logger.info(f"Letter {letter} - number {number}")
                        if not isinstance(cell, MergedCell):
                            cell.value = 'X'
                        # cont=int(number) #comvertir a entero
                        cont = cont+1
                    if diagnostic_line.observation:
                        cell2 = sheet['H'+number]
                        if not isinstance(cell2, MergedCell):
                            cell2.value = diagnostic_line.observation

        workbook.close()
        
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                xls_filelike = BytesIO(f.read())

        out = base64.encodebytes(xls_filelike.getvalue())

        self.xls_helper.write({'datas': out, 'datas_fname': filename})
        filename += '%2Exlsx'

        return {
            'type': 'ir.actions.act_url',
            'target': 'new',
            'url': 'web/content/?model='+self.xls_helper._name+'&id='+str(self.xls_helper.id)+'&field=datas&download=true&filename='+filename,
        }


    def search_all_clauseId_in_actualRequierments(self):
        fields = FIELDS
        """ Los padres de los requisitos actuales"""
        lista_clausulas = []
        for field in fields:
            for line in getattr(self, field):
                clau_id = line.clause_id
                if clau_id not in lista_clausulas:
                    lista_clausulas.append(clau_id)

        return lista_clausulas

    # ids_clause_list_old_global = all_clause.ids
    def change_state_eval(self):
        """ Aqui se verfica cada vez que se añade una clasula,
        para traer los requisitos relacionadas a esa clausula.
        Cambia el estado de draft a evalate,  cuando son seleccionadas las clausulas
        """
        if not self.all_clause:
            raise UserError("Debe seleccionar al menos una registro en ANALISIS PREVIO, antes de proceder con el análisis.")
        



        ids_clause_list_all = self.all_clause.ids

        clausulas_analis_prev = self.search_all_clauseId_in_actualRequierments()
        if (len(ids_clause_list_all)):
            ids_clause_list = []

            print("LISTA DE Existentes ", clausulas_analis_prev)
            print("LISTA DE ALL PREVIO ", clausulas_analis_prev)

            # TODO: CUANDO SE ELIMINA AUN NO FUNCIONA; FALTA QUE SEPA CUAND
            # SE ESTA ELIMINANDO el registro

            # Al inicio no hay requisitos seleccionados, por ende pasa directo
            if(len(clausulas_analis_prev)):
                for item in ids_clause_list_all:
                    # print ("ENTRO ", item)
                    if str(item) not in clausulas_analis_prev:
                        # print ("NO ESTA ", item)
                        ids_clause_list.append(item)
            else:
                ids_clause_list = ids_clause_list_all
            ids_clause_list = ids_clause_list_all


            self.diagnostic4_1_ids = self._default_diagnostic_line_ids_v2(
                '4.1', ids_clause_list)
            self.diagnostic4_2_ids = self._default_diagnostic_line_ids_v2(
                '4.2', ids_clause_list)
            self.diagnostic4_3_ids = self._default_diagnostic_line_ids_v2(
                '4.3', ids_clause_list)
            self.diagnostic4_4_ids = self._default_diagnostic_line_ids_v2(
                '4.4', ids_clause_list)
            # print("2 ---->",self.diagnostic4_ids.ids)
            self.diagnostic5_1_ids = self._default_diagnostic_line_ids_v2(
                '5.1', ids_clause_list)
            self.diagnostic5_2_ids = self._default_diagnostic_line_ids_v2(
                '5.2', ids_clause_list)
            self.diagnostic5_3_ids = self._default_diagnostic_line_ids_v2(
                '5.3', ids_clause_list)

            self.diagnostic6_1_ids = self._default_diagnostic_line_ids_v2(
                '6.1', ids_clause_list)
            self.diagnostic6_2_ids = self._default_diagnostic_line_ids_v2(
                '6.2', ids_clause_list)

            self.diagnostic7_1_ids = self._default_diagnostic_line_ids_v2(
                '7.1', ids_clause_list)
            self.diagnostic7_2_ids = self._default_diagnostic_line_ids_v2(
                '7.2', ids_clause_list)
            self.diagnostic7_3_ids = self._default_diagnostic_line_ids_v2(
                '7.3', ids_clause_list)
            self.diagnostic7_4_ids = self._default_diagnostic_line_ids_v2(
                '7.4', ids_clause_list)
            self.diagnostic7_5_ids = self._default_diagnostic_line_ids_v2(
                '7.5', ids_clause_list)

            self.diagnostic8_1_ids = self._default_diagnostic_line_ids_v2(
                '8.1', ids_clause_list)
            self.diagnostic8_2_ids = self._default_diagnostic_line_ids_v2(
                '8.2', ids_clause_list)
            self.diagnostic8_3_ids = self._default_diagnostic_line_ids_v2(
                '8.3', ids_clause_list)

            self.diagnostic9_1_ids = self._default_diagnostic_line_ids_v2(
                '9.1', ids_clause_list)
            self.diagnostic9_2_ids = self._default_diagnostic_line_ids_v2(
                '9.2', ids_clause_list)
            self.diagnostic9_3_ids = self._default_diagnostic_line_ids_v2(
                '9.3', ids_clause_list)

            self.diagnostic10_1_ids = self._default_diagnostic_line_ids_v2(
                '10.1', ids_clause_list)
            self.diagnostic10_2_ids = self._default_diagnostic_line_ids_v2(
                '10.2', ids_clause_list)
        else:
            self.diagnostic4_1_ids = False
            self.diagnostic4_2_ids = False
            self.diagnostic4_3_ids = False
            self.diagnostic4_4_ids = False

            self.diagnostic5_1_ids = False
            self.diagnostic5_2_ids = False
            self.diagnostic5_3_ids = False

            self.diagnostic6_1_ids = False
            self.diagnostic6_2_ids = False

            self.diagnostic7_1_ids = False
            self.diagnostic7_2_ids = False
            self.diagnostic7_3_ids = False
            self.diagnostic7_4_ids = False
            self.diagnostic7_5_ids = False

            self.diagnostic8_1_ids = False
            self.diagnostic8_2_ids = False
            self.diagnostic8_3_ids = False

            self.diagnostic9_1_ids = False
            self.diagnostic9_2_ids = False
            self.diagnostic9_3_ids = False
            
            self.diagnostic10_1_ids = False
            self.diagnostic10_2_ids = False

        # Cambia de estado al analizar detalladamente
        self.state = 'evaluate'

    def change_state_draft(self):
        """ Cambia el estado de evaluate a draft,
        cuando son seleccionadas las clausulas
        """
        self.state = 'draft'

    def change_state_validate(self):
        """ Cambia el estado de evaluate a validate.
        Esto culmina todo el proceso
        """
        self.state = 'validate'